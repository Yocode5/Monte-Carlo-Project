#Importing The all Important Libraries

import argparse
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import time
from openpyxl.utils import get_column_letter

#Core Simulation Logic

def estimate_pi(n):

    x_axis = np.random.uniform(-1, 1, n)
    y_axis = np.random.uniform(-1, 1, n)

    inside_circle = (x_axis**2 + y_axis**2) <= 1
    hits = inside_circle.sum()

    hit_probability = hits / n
    pi_estimate = 4 * hit_probability
    absolute_error = abs(np.pi - pi_estimate)

    return hits, hit_probability, pi_estimate, absolute_error

#Implementation of the Command line Arguments Parser

def parse_arguments():
    parser = argparse.ArgumentParser(
        description=("Monte Carlo Simulation to estimate the value of Pi using a dart-throwing method.\n"
                     "Generates Excel logs and convergence plots. \n"
                     "Example usage: python MonteCarloSim_Improved.py --n 1000 10000 100000 1000000 --runs 10 \n"
                     "Required libraries: numpy, pandas, matplotlib, openpyxl"
        )
    )

    parser.add_argument(
        "--n",
        nargs= "+",
        type=int,
        required=True,
        help="List of Dart throws (e.g. 1000 10000 100000 1000000)"
    )

    parser.add_argument(
        "--runs",
        type=int,
        default=10,
        help="Number of repetitions for each N (Default 10)"
    )

    parser.add_argument(
        "--no-plot",
        action="store_true",
        help="Disable plotting of results"
    )

    return parser.parse_args()

#Input Validation Implementation to Prevent Invalid inputs and Exceptions 

def validate_inputs(ns, runs):
    if runs <= 0:
        raise ValueError("Number of runs must be a positive integer.")
    
    for n in ns:
        if n <= 0:
            raise ValueError("All N values must be positive integers.")
        
#Main Function 
def main():
    args = parse_arguments()
    validate_inputs(args.n, args.runs)

    all_results = []

    timestamp = int(time.time())
    filename = f"MonteCarloSim_Results_{timestamp}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        for n in args.n:

            df_n = pd.DataFrame(
                 columns= [
                    "N",
                    "Run",
                    "Hits",
                    "Total_Throws",
                    "Hit_probability",
                    "Pi_Estimate",
                    "Absolute_Error",
                ]
            )

            for run in range(1, args.runs + 1):
                hits, prob, pi_est, error = estimate_pi(n)

                row = [
                    n,
                    run,
                    hits,
                    n,
                    prob,
                    pi_est,
                    error,
                ]

                df_n.loc[len(df_n)] = row
                all_results.append(row)

            sheet_name = f"N_{n}"
            df_n.to_excel(writer, sheet_name=sheet_name, index=False)

            Worksheet = writer.sheets[sheet_name]
            Worksheet.freeze_panes = "A2"

            for i, col in enumerate(Worksheet.columns, start=1):
                max_length = 0
                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                Worksheet.column_dimensions[get_column_letter(i)].width = max_length + 2


        #Summary Statistics

        df_all = pd.DataFrame(
            all_results,
            columns=[
                "N",
                "Run",
                "Hits",
                "Total_Throws",
                "Hit_probability",
                "Pi_Estimate",
                "Absolute_Error"
            ]
        )

        summary_df = df_all.groupby("N").agg(
            Mean_Pi = ("Pi_Estimate", "mean"),
            Mode_Pi = ("Pi_Estimate", lambda x: x.mode().iloc[0]),
            Mean_Error = ("Absolute_Error", "mean"),
        ).reset_index()

        summary_df.to_excel(
            writer, sheet_name="Summary_Statistics", index=False
        )

        Worksheet = writer.sheets["Summary_Statistics"]
        Worksheet.freeze_panes = "A2"

        for i, col in enumerate(Worksheet.columns, start=1):
            max_length = 0
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            Worksheet.column_dimensions[get_column_letter(i)].width = max_length + 2 

    print(f"\nExcel Results saved to {filename}")

    if not args.no_plot:
        plt.figure(figsize=(8, 5))
        plt.plot(
            summary_df["N"],
            summary_df["Mean_Pi"],
            marker="o",
            linewidth=2,
            label="Mean Pi Estimate",
        )

        plt.axhline(
            np.pi,
            linestyle="--",
            linewidth=1.8,
            color="red",
            label="Actual Pi Value",
        )

        plt.xscale("log")
        plt.xlabel("Number of Dart Throws (N)")
        plt.ylabel("Estimated Value of Pi")
        plt.title("Monte Carlo Simulation for Estimating Pi")
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.show()

if __name__ == "__main__":
    main()


